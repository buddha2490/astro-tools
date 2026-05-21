#!/usr/bin/env python3
"""
Extract object lists from the three astrophotography target workbooks in
data-raw/, enrich each object with type / magnitude / size from SIMBAD, derive
the constellation from RA/Dec, and write one JSON file per workbook.

Coordinates come from the workbooks (the user's source of truth); SIMBAD is used
only for type, magnitude, size, and to confirm resolution. Constellation is
derived from the workbook RA/Dec via astropy.

Run with the project venv:
    /tmp/catimport-venv/bin/python extract_enrich.py
"""
import json
import os
import re
import sys

import openpyxl
from astropy import units as u
from astropy.coordinates import SkyCoord, get_constellation
from astroquery.simbad import Simbad

HERE = os.path.dirname(os.path.abspath(__file__))
DATA_RAW = os.path.normpath(os.path.join(HERE, "..", "..", "data-raw"))
OUT_DIR = HERE

# --- which workbook -> table, and where the object columns live -------------
# Each Visibility tab shows the object list twice (two side-by-side panels).
# We read both panels and dedupe by designation. Column letters are the
# (designation, common_name, ra_hours, dec_deg) for each panel.
#
# Dark Nebula has two sheets instead of one "Visibility" tab; use the `sheets`
# key as a list of (sheet_name, panels) pairs in that case.
WORKBOOKS = [
    {
        "file": "Messier 110 Objects V1.2.xlsx",
        "table": "catalog_messier",
        "panels": [("F", "J", "L", "M", "K"), ("Q", "U", "W", "X", "V")],
        "has_difficulty": True,  # 5th column letter = Difficulty
    },
    {
        "file": "NEW! Top 100 Astrophotography Objects V1.0.xlsx",
        "table": "catalog_top100",
        "panels": [("J", "M", "N", "O", None), ("U", "X", "Y", "Z", None)],
        "has_difficulty": False,
    },
    {
        "file": "Nebulae 100 Objects V1.2.xlsx",
        "table": "catalog_nebulae",
        "panels": [("H", "K", "L", "M", None), ("S", "V", "W", "X", None)],
        "has_difficulty": False,
    },
    {
        "file": "Dark Nebula V1.2.xlsx",
        "table": "catalog_dark_nebulae",
        # Two sheets, each with two side-by-side panels.
        # Panels: (designation, common_name/notes, ra_hours, dec_deg, difficulty)
        # Notes column used as common_name; numeric-only values are discarded.
        "sheets": [
            ("Visibility (Required)", [
                ("F", "L", "R", "S", None),    # left: Name, Notes, RA, Decl
                ("X", "AD", "AJ", "AK", None), # right: Name, Notes, RA, Decl
            ]),
            ("Visibility (Optional)", [
                ("F", "L", "Q", "R", None),    # left: Name, Notes, RA, Decl
                ("V", "AB", "AG", "AH", None), # right: Name, Notes, RA, Decl
            ]),
        ],
        "has_difficulty": False,
        "data_start_row": 9,  # row 8 is a header-continuation row in this workbook
    },
]

DATA_START_ROW = 8

# Categories that are NOT overridden by a "...Nebula" common name: galaxies
# (e.g. M74 "Phantom ...Galaxy") and types that are already a nebula.
NEBULA_TYPES_KEEP = {
    "Galaxy",
    "Planetary Nebula",
    "Supernova Remnant",
    "HII Region (Emission Nebula)",
    "Reflection Nebula",
    "Dark Nebula",
    "Nebula",
    "Molecular Cloud",
    "Emission/Reflection Nebula",
}

# Workbook designations that SIMBAD knows under a different identifier.
# Key = designation as written in the workbook; value = SIMBAD-resolvable name.
# The workbook designation is still what we store; the alias is only for lookup.
SIMBAD_ALIASES = {
    "Jones-Emberson 1": "PK 164+31.1",  # = PN ARO 121
    "NGC 650/1": "NGC 650",             # M76, Little Dumbbell (NGC 650+651)
    "IC 5148/50": "IC 5148",            # Spare Tyre Nebula
    "Shapley 1": "PN Sp 1",
    "Vela Pencil": "NGC 2736",          # Pencil Nebula
    "Abell 21": "A66 21",               # Medusa Nebula (PN); bare "Abell 21"
                                        # resolves to a galaxy cluster in SIMBAD
    # Not in SIMBAD under any common name (informal / very faint targets):
    #   "Ou 4" (Giant Squid), "Vela Bridge", "Vela Spiral Flame"
}


def extract_objects(path, panels, has_difficulty, sheet_name="Visibility",
                    data_start_row=DATA_START_ROW):
    """Return {designation: {...}} deduped across both display panels."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[sheet_name]
    max_row = ws.max_row

    def cell(col, row):
        return ws[f"{col}{row}"].value
    grid = {}
    for desig_c, name_c, ra_c, dec_c, diff_c in panels:
        for col in (desig_c, name_c, ra_c, dec_c, diff_c):
            if col is None:
                continue
            for r in range(data_start_row, max_row + 1):
                grid[f"{col}{r}"] = cell(col, r)
    wb.close()

    out = {}
    for desig_c, name_c, ra_c, dec_c, diff_c in panels:
        for r in range(data_start_row, max_row + 1):
            desig = grid.get(f"{desig_c}{r}")
            if desig is None or str(desig).strip() == "":
                continue
            # normalize internal whitespace (e.g. "B  34" -> "B 34")
            desig = " ".join(str(desig).split())
            name = grid.get(f"{name_c}{r}")
            # discard None, empty strings, and numeric-only notes (workbook
            # formula cells return 0 for empty rows in some sheets)
            if name is None or name == "" or isinstance(name, (int, float)):
                name = None
            else:
                name = str(name).strip() or None
            ra = grid.get(f"{ra_c}{r}")
            dec = grid.get(f"{dec_c}{r}")
            diff = grid.get(f"{diff_c}{r}") if (has_difficulty and diff_c) else None
            diff = str(diff).strip() if diff not in (None, "") else None

            rec = out.setdefault(desig, {})
            # prefer first seen numeric coords; fill name/difficulty if missing
            if "ra_hours" not in rec and isinstance(ra, (int, float)):
                rec["ra_hours"] = float(ra)
            if "dec_deg" not in rec and isinstance(dec, (int, float)):
                rec["dec_deg"] = float(dec)
            if not rec.get("common_name") and name:
                rec["common_name"] = name
            if not rec.get("difficulty") and diff:
                rec["difficulty"] = diff
            rec.setdefault("designation", desig)
    return out


# --- SIMBAD otype code -> coarse, astrophotographer-friendly category --------
def build_otype_category_map(simbad):
    """Map every SIMBAD otype code to a coarse category using the otype label."""
    defs = simbad.query_tap("SELECT otype, label, description FROM otypedef")
    label_by_code = {row["otype"]: row["label"] for row in defs}

    galaxy_codes = set()
    for code, label in label_by_code.items():
        lab = label.lower()
        if ("galax" in lab or code in {
            "G", "GiG", "GiC", "GiP", "BiC", "IG", "PaG", "GrG", "ClG", "SCG",
            "AGN", "SyG", "Sy1", "Sy2", "LIN", "QSO", "Bla", "BLL", "rG", "brG",
            "EmG", "SBG", "H2G", "LSB", "Q?", "EmO",
        }) and code not in {"EmO"}:
            galaxy_codes.add(code)

    def categorize(code):
        if code is None:
            return None
        label = label_by_code.get(code, code)
        if code in galaxy_codes:
            return "Galaxy"
        coarse = {
            "GlC": "Globular Cluster",
            "OpC": "Open Cluster",
            "Cl*": "Open Cluster",
            "As*": "Stellar Association",
            "PN": "Planetary Nebula",
            "PN?": "Planetary Nebula",
            "SNR": "Supernova Remnant",
            "SR?": "Supernova Remnant",
            "HII": "HII Region (Emission Nebula)",
            "RNe": "Reflection Nebula",
            "DNe": "Dark Nebula",
            "GNe": "Nebula",
            "ISM": "Nebula",
            "Cld": "Nebula",
            "MoC": "Molecular Cloud",
            "SFR": "Star Forming Region",
            "bub": "Nebula",
            "EmO": "Emission Object",
            "**": "Double Star",
            "*": "Star",
            "V*": "Variable Star",
        }
        return coarse.get(code, label)

    return label_by_code, categorize


def _simbad_query_name(desig):
    """Return the SIMBAD-resolvable name for a designation."""
    if desig in SIMBAD_ALIASES:
        return SIMBAD_ALIASES[desig]
    # Barnard dark nebula: "B NNN" -> "Barnard NNN"
    m = re.match(r'^B\s+(\d+)$', desig)
    if m:
        return f"Barnard {m.group(1)}"
    # Sandqvist-Lindroos: "SL NNN" -> "Sandqvist NNN"
    m = re.match(r'^SL\s+(\d+)$', desig)
    if m:
        return f"Sandqvist {m.group(1)}"
    return desig


def enrich(simbad, designations):
    """Query SIMBAD for a batch of designations. Returns {desig: {...}}."""
    simbad.reset_votable_fields()
    simbad.add_votable_fields(
        "otype", "otype_txt", "galdim_majaxis", "galdim_minaxis", "V", "B"
    )
    # query under alias when one exists; map results back to the designation
    query_names = [_simbad_query_name(d) for d in designations]
    desig_by_query = {_simbad_query_name(d): d for d in designations}
    table = simbad.query_objects(query_names)
    res = {}
    if table is None:
        return res
    for row in table:
        key = row["user_specified_id"]
        if key is None:
            continue
        key = key.strip() if isinstance(key, str) else str(key).strip()
        key = desig_by_query.get(key, key)

        def num(col):
            v = row[col]
            try:
                if v is None or (hasattr(v, "mask") and v is None):
                    return None
                fv = float(v)
                return fv if fv == fv else None  # drop NaN
            except (TypeError, ValueError):
                return None

        main_id = row["main_id"]
        main_id = main_id.strip() if isinstance(main_id, str) else (
            main_id.decode() if isinstance(main_id, bytes) else None
        )
        otype = row["otype"]
        otype = otype.strip() if isinstance(otype, str) else (
            otype.decode().strip() if isinstance(otype, bytes) else None
        )
        v_mag = num("V")
        b_mag = num("B")
        if v_mag is not None:
            mag, band = round(v_mag, 2), "V"
        elif b_mag is not None:
            mag, band = round(b_mag, 2), "B"
        else:
            mag, band = None, None
        maj = num("galdim_majaxis")
        minr = num("galdim_minaxis")
        res[str(key)] = {
            "simbad_main_id": main_id or None,
            "simbad_otype": otype or None,
            "magnitude": mag,
            "magnitude_band": band,
            "size_major_arcmin": round(maj, 3) if maj is not None else None,
            "size_minor_arcmin": round(minr, 3) if minr is not None else None,
        }
    return res


def main():
    simbad = Simbad(ROW_LIMIT=-1)
    simbad.TIMEOUT = 120
    label_by_code, categorize = build_otype_category_map(simbad)

    summary = []
    for wb_cfg in WORKBOOKS:
        path = os.path.join(DATA_RAW, wb_cfg["file"])
        data_start_row = wb_cfg.get("data_start_row", DATA_START_ROW)
        if "sheets" in wb_cfg:
            objs = {}
            for sheet_name, panels in wb_cfg["sheets"]:
                sheet_objs = extract_objects(
                    path, panels, wb_cfg["has_difficulty"],
                    sheet_name=sheet_name, data_start_row=data_start_row,
                )
                for desig, rec in sheet_objs.items():
                    objs.setdefault(desig, rec)  # keep first-seen
        else:
            objs = extract_objects(path, wb_cfg["panels"], wb_cfg["has_difficulty"],
                                   data_start_row=data_start_row)
        designations = list(objs.keys())
        enriched = enrich(simbad, designations)

        rows = []
        unresolved = []
        for desig in designations:
            o = objs[desig]
            e = enriched.get(desig, {})
            ra_hours = o.get("ra_hours")
            dec_deg = o.get("dec_deg")
            constellation = None
            ra_deg = None
            if ra_hours is not None and dec_deg is not None:
                ra_deg = round(ra_hours * 15.0, 6)
                coord = SkyCoord(ra=ra_deg * u.deg, dec=dec_deg * u.deg,
                                 frame="icrs")
                constellation = get_constellation(coord)

            otype = e.get("simbad_otype")
            if not e.get("simbad_main_id"):
                unresolved.append(desig)

            object_type = categorize(otype) if otype else None
            # Name-based override: many emission/reflection nebulae share a
            # catalog number with their embedded star cluster, so SIMBAD types
            # the identifier as a cluster/star/YSO/WR/radio object. When the
            # common name says "...Nebula", trust that for object_type (raw
            # simbad_otype is preserved). Galaxies and already-nebula types are
            # left untouched.
            cn = (o.get("common_name") or "").lower()
            # Only override when SIMBAD returned *something* that is wrong; if
            # SIMBAD returned nothing (object_type is None), leave it as-is so
            # we don't misclassify unresolved dark nebulae as emission nebulae.
            if "nebula" in cn and object_type is not None and object_type not in NEBULA_TYPES_KEEP:
                object_type = "Emission/Reflection Nebula"

            maj = e.get("size_major_arcmin")
            minr = e.get("size_minor_arcmin")
            if maj is not None and minr is not None:
                size_txt = f"{maj:g} x {minr:g}"
            elif maj is not None:
                size_txt = f"{maj:g}"
            else:
                size_txt = None

            rows.append({
                "designation": desig,
                "common_name": o.get("common_name"),
                "object_type": object_type,
                "simbad_otype": otype,
                "simbad_otype_label": label_by_code.get(otype) if otype else None,
                "ra_hours": ra_hours,
                "ra_deg": ra_deg,
                "dec_deg": dec_deg,
                "constellation": constellation,
                "magnitude": e.get("magnitude"),
                "magnitude_band": e.get("magnitude_band"),
                "size_major_arcmin": maj,
                "size_minor_arcmin": minr,
                "size_arcmin": size_txt,
                "simbad_main_id": e.get("simbad_main_id"),
                **({"difficulty": o.get("difficulty")} if wb_cfg["has_difficulty"] else {}),
            })

        out_path = os.path.join(OUT_DIR, wb_cfg["table"] + ".json")
        with open(out_path, "w") as fh:
            json.dump({"table": wb_cfg["table"],
                       "source_file": wb_cfg["file"],
                       "rows": rows}, fh, indent=2)
        summary.append((wb_cfg["table"], len(rows), unresolved))
        print(f"{wb_cfg['table']}: {len(rows)} objects -> {out_path}")
        if unresolved:
            print(f"   UNRESOLVED in SIMBAD ({len(unresolved)}): {unresolved}")

    print("\nDone.")
    return summary


if __name__ == "__main__":
    main()
